"""Prévisualisation et export Excel avec responsables ecclésiaux dynamiques."""

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET

from ..models import District, FicheParoisse, Profil, Province, Region, Zone
from ..permissions import get_role, role_required
from ..services.services_responsables_ecclesiaux import (
    construire_index_responsables,
    responsables_pour_fiche,
)
from .helpers import _fiches_visibles_pour


def _fiches_export_filtrees(request):
    fiches = _fiches_visibles_pour(request.user)
    role = get_role(request.user)
    statut_filtre = request.GET.get("statut", "")
    if role == Profil.Role.SUPER_ADMIN:
        if statut_filtre == "attente_superviseur":
            fiches = fiches.filter(statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR)
        elif statut_filtre == "attente_manager":
            fiches = fiches.filter(statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER)
        elif statut_filtre != "tous":
            fiches = fiches.filter(statut_validation=FicheParoisse.StatutValidation.VALIDEE)

    def valid_id(name):
        value = (request.GET.get(name) or "").strip()
        return int(value) if value.isdigit() else None

    region_id, province_id = valid_id("region"), valid_id("province")
    district_id, zone_id = valid_id("district"), valid_id("zone")
    paroisse = (request.GET.get("paroisse") or "").strip()[:100]
    if region_id:
        fiches = fiches.filter(region_id=region_id)
    if province_id:
        fiches = fiches.filter(province_id=province_id)
    if district_id:
        fiches = fiches.filter(district_id=district_id)
    if zone_id:
        fiches = fiches.filter(zone_id=zone_id)
    if paroisse:
        fiches = fiches.filter(nom_paroisse__icontains=paroisse)

    return fiches.select_related("region", "province", "district", "zone", "village").order_by(
        "region__nom", "province__nom", "district__nom", "zone__nom", "nom_paroisse"
    )


def _fiches_avec_responsables(request):
    fiches, index = construire_index_responsables(_fiches_export_filtrees(request))
    for fiche in fiches:
        fiche.responsables_ecclesiaux = responsables_pour_fiche(fiche, index)
    return fiches


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def fiche_export_preview(request):
    fiches = _fiches_avec_responsables(request)
    hierarchy = {}
    for fiche in fiches:
        hierarchy.setdefault(fiche.region.nom, {}).setdefault(fiche.province.nom, {}).setdefault(
            fiche.district.nom, {}
        ).setdefault(fiche.zone.nom, []).append(fiche)

    def selected(model, name):
        value = request.GET.get(name, "")
        return model.objects.filter(pk=value).first() if value.isdigit() else None

    region = selected(Region, "region")
    province = selected(Province, "province")
    district = selected(District, "district")
    zone = selected(Zone, "zone")
    return render(
        request,
        "recensement/fiche_export_preview.html",
        {
            "hierarchy": hierarchy,
            "total": len(fiches),
            "filters": {
                "statut": request.GET.get("statut", ""),
                "region": region.nom if region else "",
                "province": province.nom if province else "",
                "district": district.nom if district else "",
                "zone": zone.nom if zone else "",
                "paroisse": request.GET.get("paroisse", ""),
            },
            "query_string": request.GET.urlencode(),
        },
    )


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def fiche_export_excel(request):
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fiches = _fiches_avec_responsables(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Paroisses"
    headers = [
        "Code officiel",
        "Région",
        "Titre responsable région",
        "Nom responsable région",
        "Province",
        "Titre responsable province",
        "Nom responsable province",
        "District",
        "Titre responsable district",
        "Nom responsable district",
        "Zone",
        "Titre responsable zone",
        "Nom responsable zone",
        "Paroisse",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(*[Side(style="thin", color="E5E7EB")] * 4)
    for cell in ws[1]:
        cell.fill, cell.font, cell.border = header_fill, header_font, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for fiche in fiches:
        r = fiche.responsables_ecclesiaux
        ws.append(
            [
                fiche.code_officiel or "Code officiel en attente",
                fiche.region.nom,
                r["region"]["titre"],
                r["region"]["nom"],
                fiche.province.nom,
                r["province"]["titre"],
                r["province"]["nom"],
                fiche.district.nom,
                r["district"]["titre"],
                r["district"]["nom"],
                fiche.zone.nom,
                r["zone"]["titre"],
                r["zone"]["nom"],
                fiche.nom_paroisse or "",
            ]
        )
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [34, 24, 30, 28, 28, 30, 28, 30, 30, 28, 32, 30, 28, 40]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    recap = wb.create_sheet("Synthèse")
    recap["A1"] = "Export hiérarchique des paroisses et responsables ecclésiaux"
    recap["A1"].font = Font(bold=True, size=14)
    recap["A3"], recap["B3"] = "Nombre de paroisses concernées", len(fiches)
    recap["A5"], recap["B5"] = "Source des responsables", "Module autonome des postes et mandats ecclésiaux"
    recap["A7"], recap["B7"] = "Valeur en cas d'absence", "Non renseigné"
    recap.column_dimensions["A"].width, recap.column_dimensions["B"].width = 34, 90
    for row in recap.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="paroisses_responsables_ecclesiaux.xlsx"'
    return response
