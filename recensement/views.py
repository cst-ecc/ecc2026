import csv
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_http_methods

from .forms import (
    FicheParoisseForm, MotifModificationForm, PhotosParoisseForm, ProfilForm,
    TailwindSetPasswordForm,
)
from .identifiants import generer_identifiant, generer_mot_de_passe_provisoire
from .codification import generer_code_paroisse
from .models import (
    District, FicheParoisse, HistoriqueModification, PhotoParoisse, Profil,
    Province, Region, Village, Zone,
)
from .permissions import (
    districts_autorises, fiche_dans_perimetre, fiches_visibles_pour, get_role, peut_creer_utilisateur,
    peut_modifier_fiche, peut_valider_fiche, perimetre_creation_autorise,
    role_required, roles_creables_par, zones_autorisees,
)

# Caractères qu'un tableur peut interpréter comme début de formule (OWASP CSV Injection).
_CSV_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value):
    text = "" if value is None else str(value)
    if text.startswith(_CSV_FORMULA_PREFIXES):
        return "'" + text
    return text


# Associe chaque champ du wizard à son étape (index JS, base 0).
_CHAMP_VERS_ETAPE = {
    "region": 0, "province": 0, "district": 0, "zone": 0, "village": 0,
    "nouvelle_localite_nom": 0,
    "nom_paroisse": 1, "annee_fondation": 1, "statut_batiment": 1, "nombre_fideles_estime": 1,
    "photos": 1,
    "parish_shepherd": 2, "contact_responsable": 2, "photo_charge": 2,
    "latitude": 3, "longitude": 3, "precision_gps": 3, "observations": 3,
    "nom_informateur": 4, "contact_informateur": 4,
}


def _premiere_etape_en_erreur(form, photos_form=None):
    etapes = set()
    for champ in form.errors:
        etapes.add(_CHAMP_VERS_ETAPE.get(champ, 0))
    if form.non_field_errors():
        etapes.add(0)
    if photos_form is not None and photos_form.errors:
        etapes.add(_CHAMP_VERS_ETAPE.get("photos", 1))
    return min(etapes) if etapes else None


def _snapshot_fiche(fiche):
    return {
        "region": fiche.region.nom,
        "province": fiche.province.nom,
        "district": fiche.district.nom,
        "zone": fiche.zone.nom,
        "village": fiche.village.nom if fiche.village_id else None,
        "nouvelle_localite_nom": fiche.nouvelle_localite_nom,
        "nom_paroisse": fiche.nom_paroisse,
        "annee_fondation": fiche.annee_fondation,
        "parish_shepherd": fiche.parish_shepherd,
        "contact_responsable": fiche.contact_responsable,
        "photo_charge": fiche.photo_charge.name if fiche.photo_charge else None,
        "nombre_fideles_estime": fiche.nombre_fideles_estime,
        "statut_batiment": fiche.get_statut_batiment_display(),
        "latitude": str(fiche.latitude) if fiche.latitude is not None else None,
        "longitude": str(fiche.longitude) if fiche.longitude is not None else None,
        "precision_gps": fiche.precision_gps,
        "nom_informateur": fiche.nom_informateur,
        "contact_informateur": fiche.contact_informateur,
        "observations": fiche.observations,
    }


def _fiches_visibles_pour(user):
    """Compatibilité locale : délègue au moteur territorial centralisé."""
    return fiches_visibles_pour(user)


# ---------------------------------------------------------------------------
# Pages publiques / aiguillage
# ---------------------------------------------------------------------------

def landing(request):
    if request.user.is_authenticated:
        return redirect("recensement:post_login_redirect")
    return render(request, "recensement/landing.html")


@login_required
def post_login_redirect(request):
    """Aiguillage après connexion selon le rôle."""
    if get_role(request.user) == Profil.Role.SUPER_ADMIN:
        return redirect("recensement:dashboard")
    return redirect("recensement:fiche_list")


# ---------------------------------------------------------------------------
# Fiches de recensement
# ---------------------------------------------------------------------------

@login_required
@role_required(Profil.Role.AGENT, Profil.Role.SUPER_ADMIN)
@require_http_methods(["GET", "POST"])
def fiche_create(request):
    """Formulaire de saisie terrain. Réservé aux agents et au super admin."""
    etape_erreur = None
    if request.method == "POST":
        form = FicheParoisseForm(request.POST, request.FILES, user=request.user)
        photos_form = PhotosParoisseForm(request.POST, request.FILES)
        if form.is_valid() and photos_form.is_valid():
            fiche = form.save(commit=False)
            fiche.cree_par = request.user
            fiche.save()
            for photo in photos_form.cleaned_data["photos"]:
                PhotoParoisse.objects.create(fiche=fiche, image=photo)
            messages.success(
                request,
                "Fiche enregistrée avec succès, en attente de validation par l'OP DISTRICT. "
                "Vous pouvez recenser une autre paroisse.",
            )
            return redirect("recensement:fiche_create")
        etape_erreur = _premiere_etape_en_erreur(form, photos_form)
        messages.error(
            request,
            "La fiche n'a pas pu être enregistrée : le formulaire s'est rouvert directement "
            "à l'étape contenant l'erreur, indiquée en rouge ci-dessous.",
        )
    else:
        form = FicheParoisseForm(user=request.user)
        photos_form = PhotosParoisseForm()

    initial_ids = {
        "region": form["region"].value(),
        "province": form["province"].value(),
        "district": form["district"].value(),
        "zone": form["zone"].value(),
        "village": form["village"].value(),
    }

    return render(request, "recensement/fiche_form.html", {
        "form": form,
        "photos_form": photos_form,
        "initial_ids_json": json.dumps(initial_ids),
        "etape_erreur_json": json.dumps(etape_erreur),
        "current_role": get_role(request.user),
        "is_super_admin": get_role(request.user) == Profil.Role.SUPER_ADMIN,
    })


@login_required
@role_required(Profil.Role.OP_DISTRICT, Profil.Role.OP_PROVINCE)
@require_http_methods(["GET", "POST"])
def fiche_update(request, pk):
    """Modification d'une fiche — réservée à l'OP DISTRICT (son district) et
    à l'OP PROVINCE (sa province), selon le palier de validation en cours."""
    fiche = get_object_or_404(FicheParoisse, pk=pk)

    if not peut_modifier_fiche(request.user, fiche):
        role = get_role(request.user)
        profil = getattr(request.user, "profil", None)
        hors_perimetre = not fiche_dans_perimetre(request.user, fiche)
        if hors_perimetre:
            messages.error(request, "Cette fiche n'est pas dans votre périmètre (district/province).")
            return redirect("recensement:fiche_list")
        else:
            messages.error(
                request,
                "Cette fiche a déjà été validée à votre niveau et ne peut plus être "
                "modifiée — elle relève désormais du palier suivant.",
            )
            return redirect("recensement:fiche_detail", pk=fiche.pk)

    if request.method == "POST":
        form = FicheParoisseForm(request.POST, request.FILES, instance=fiche, user=request.user)
        motif_form = MotifModificationForm(request.POST)
        if form.is_valid() and motif_form.is_valid():
            avant = _snapshot_fiche(fiche)
            fiche_modifiee = form.save()
            apres = _snapshot_fiche(fiche_modifiee)
            HistoriqueModification.objects.create(
                fiche=fiche_modifiee,
                modifie_par=request.user,
                motif=motif_form.cleaned_data["motif"],
                donnees_avant=avant,
                donnees_apres=apres,
            )
            messages.success(
                request,
                "Fiche modifiée avec succès. Le motif a été enregistré dans l'historique.",
            )
            return redirect("recensement:fiche_detail", pk=fiche.pk)
        messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = FicheParoisseForm(instance=fiche, user=request.user)
        motif_form = MotifModificationForm()

    initial_ids = {
        "region": fiche.region_id,
        "province": fiche.province_id,
        "district": fiche.district_id,
        "zone": fiche.zone_id,
        "village": fiche.village_id,
    }

    return render(request, "recensement/fiche_edit_form.html", {
        "form": form,
        "motif_form": motif_form,
        "fiche": fiche,
        "initial_ids_json": json.dumps(initial_ids),
        "current_role": get_role(request.user),
        "is_super_admin": get_role(request.user) == Profil.Role.SUPER_ADMIN,
    })


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_http_methods(["GET", "POST"])
def fiche_delete(request, pk):
    fiche = get_object_or_404(FicheParoisse, pk=pk)
    if request.method == "POST":
        nom = fiche.nom_paroisse
        fiche.delete()
        messages.success(request, f"La fiche « {nom} » a été supprimée définitivement.")
        return redirect("recensement:fiche_list")
    return render(request, "recensement/fiche_confirm_delete.html", {"fiche": fiche})


@login_required
@require_GET
def fiche_list(request):
    """Liste des fiches filtrée selon le rôle.

    - Super admin : filtres complets (statut, hiérarchie, paroisse).
    - Autres rôles : limité par _fiches_visibles_pour().
    """
    fiches = _fiches_visibles_pour(request.user)
    role = get_role(request.user)

    statut_filtre = ""
    region_id = None
    province_id = None
    district_id = None
    zone_id = None
    paroisse = ""

    regions = Region.objects.none()
    provinces = Province.objects.none()
    districts = District.objects.none()
    zones = Zone.objects.none()
    paroisses = []

    if role == Profil.Role.SUPER_ADMIN:
        statut_filtre = request.GET.get("statut", "")

        if statut_filtre == "attente_superviseur":
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR
            )
        elif statut_filtre == "attente_manager":
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER
            )
        elif statut_filtre == "tous":
            pass
        else:
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.VALIDEE
            )
            statut_filtre = "validees"

        def get_valid_id(param_name, model):
            value = (request.GET.get(param_name) or "").strip()
            if value.isdigit() and model.objects.filter(pk=int(value)).exists():
                return int(value)
            return None

        region_id = get_valid_id("region", Region)
        province_id = get_valid_id("province", Province)
        district_id = get_valid_id("district", District)
        zone_id = get_valid_id("zone", Zone)
        paroisse = (request.GET.get("paroisse") or "").strip()[:100]

        if region_id:
            fiches = fiches.filter(region_id=region_id)
        if province_id:
            fiches = fiches.filter(province_id=province_id)
        if district_id:
            fiches = fiches.filter(district_id=district_id)
        if zone_id:
            fiches = fiches.filter(zone_id=zone_id)
        if paroisse:
            fiches = fiches.filter(nom_paroisse__icontains=paroisse)

        regions = Region.objects.all().order_by("nom")
        provinces = Province.objects.all().order_by("nom")
        if region_id:
            provinces = provinces.filter(region_id=region_id)

        districts = District.objects.all().order_by("nom")
        if province_id:
            districts = districts.filter(province_id=province_id)
        elif region_id:
            districts = districts.filter(province__region_id=region_id)

        zones = Zone.objects.all().order_by("nom")
        if district_id:
            zones = zones.filter(district_id=district_id)
        elif province_id:
            zones = zones.filter(district__province_id=province_id)
        elif region_id:
            zones = zones.filter(district__province__region_id=region_id)

        paroisses_qs = _fiches_visibles_pour(request.user)
        if statut_filtre == "attente_superviseur":
            paroisses_qs = paroisses_qs.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR
            )
        elif statut_filtre == "attente_manager":
            paroisses_qs = paroisses_qs.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER
            )
        elif statut_filtre == "tous":
            pass
        else:
            paroisses_qs = paroisses_qs.filter(
                statut_validation=FicheParoisse.StatutValidation.VALIDEE
            )

        if zone_id:
            paroisses_qs = paroisses_qs.filter(zone_id=zone_id)
        elif district_id:
            paroisses_qs = paroisses_qs.filter(district_id=district_id)
        elif province_id:
            paroisses_qs = paroisses_qs.filter(province_id=province_id)
        elif region_id:
            paroisses_qs = paroisses_qs.filter(region_id=region_id)

        paroisses = (
            paroisses_qs
            .order_by("nom_paroisse")
            .values_list("nom_paroisse", flat=True)
            .distinct()
        )

    context = {
        "fiches": fiches.select_related(
            "region", "province", "district", "zone", "cree_par"
        )[:500],
        "regions": regions,
        "provinces": provinces,
        "districts": districts,
        "zones": zones,
        "paroisses": paroisses,
        "region_id": region_id,
        "province_id": province_id,
        "district_id": district_id,
        "zone_id": zone_id,
        "paroisse": paroisse,
        "total": fiches.count(),
        "statut_filtre": statut_filtre,
    }

    return render(request, "recensement/fiche_list.html", context)


@login_required
@require_GET
def fiche_detail(request, pk):
    """Détail d'une fiche — 404 si hors du périmètre visible (anti-IDOR)."""
    fiche = get_object_or_404(_fiches_visibles_pour(request.user), pk=pk)
    context = {
        "fiche": fiche,
        "peut_modifier": peut_modifier_fiche(request.user, fiche),
        "peut_valider": peut_valider_fiche(request.user, fiche),
    }
    if get_role(request.user) == Profil.Role.SUPER_ADMIN:
        context["historique"] = fiche.historique.select_related("modifie_par")
    return render(request, "recensement/fiche_detail.html", context)


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def fiche_export_preview(request):
    fiches = _fiches_export_filtrees(request)
    total = fiches.count()
    hierarchy = {}
    for fiche in fiches:
        region_nom = fiche.region.nom if fiche.region else "—"
        province_nom = fiche.province.nom if fiche.province else "—"
        district_nom = fiche.district.nom if fiche.district else "—"
        zone_nom = fiche.zone.nom if fiche.zone else "—"
        hierarchy.setdefault(region_nom, {})
        hierarchy[region_nom].setdefault(province_nom, {})
        hierarchy[region_nom][province_nom].setdefault(district_nom, {})
        hierarchy[region_nom][province_nom][district_nom].setdefault(zone_nom, [])
        hierarchy[region_nom][province_nom][district_nom][zone_nom].append(fiche)

    region = Region.objects.filter(pk=request.GET.get("region")).first() if request.GET.get("region", "").isdigit() else None
    province = Province.objects.filter(pk=request.GET.get("province")).first() if request.GET.get("province", "").isdigit() else None
    district = District.objects.filter(pk=request.GET.get("district")).first() if request.GET.get("district", "").isdigit() else None
    zone = Zone.objects.filter(pk=request.GET.get("zone")).first() if request.GET.get("zone", "").isdigit() else None

    filters = {
        "statut": request.GET.get("statut", ""),
        "region": region.nom if region else "",
        "province": province.nom if province else "",
        "district": district.nom if district else "",
        "zone": zone.nom if zone else "",
        "paroisse": request.GET.get("paroisse", ""),
    }

    return render(request, "recensement/fiche_export_preview.html", {
        "hierarchy": hierarchy,
        "total": total,
        "filters": filters,
        "query_string": request.GET.urlencode(),
    })


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def fiche_export_excel(request):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fiches = _fiches_export_filtrees(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Paroisses"

    headers = ["Code officiel", "Région", "Province", "District", "Zone", "Paroisse"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for fiche in fiches:
        ws.append([
            fiche.code_officiel or "Code officiel en attente",
            fiche.region.nom if fiche.region else "",
            fiche.province.nom if fiche.province else "",
            fiche.district.nom if fiche.district else "",
            fiche.zone.nom if fiche.zone else "",
            fiche.nom_paroisse or "",
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {"A": 34, "B": 24, "C": 28, "D": 30, "E": 32, "F": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    recap = wb.create_sheet("Synthèse")
    recap["A1"] = "Prévisualisation de l'export"
    recap["A1"].font = Font(bold=True, size=14)
    recap["A3"] = "Nombre de paroisses concernées"
    recap["B3"] = fiches.count()
    recap["A5"] = "Organisation des colonnes"
    recap["B5"] = "Code officiel → Région → Province → District → Zone → Paroisse"
    recap["A7"] = "Colonnes exclues"
    recap["B7"] = "Statut du bâtiment, GPS, Agent, Statut, Date, Actions"
    recap.column_dimensions["A"].width = 32
    recap.column_dimensions["B"].width = 80

    for row in recap.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="paroisses_hierarchie.xlsx"'
    return response


# ---------------------------------------------------------------------------
# Workflow de validation hiérarchique
# ---------------------------------------------------------------------------

@login_required
@role_required(Profil.Role.OP_DISTRICT, Profil.Role.OP_PROVINCE)
@require_GET
def fiche_a_valider(request):
    """File d'attente de validation selon le rôle connecté."""
    role = get_role(request.user)
    profil = getattr(request.user, "profil", None)

    if role == Profil.Role.OP_DISTRICT:
        zone_ids = zones_autorisees(request.user) or set()
        fiches = FicheParoisse.objects.filter(
            statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR,
            zone_id__in=zone_ids,
        )
    else:  # OP_PROVINCE
        fiches = FicheParoisse.objects.filter(
            statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER,
            province_id=profil.province_id if profil else None,
        )

    fiches = fiches.select_related(
        "region", "province", "district", "zone", "village", "cree_par"
    ).order_by("date_recensement")

    return render(request, "recensement/fiche_a_valider.html", {"fiches": fiches, "role": role})


@login_required
@role_required(Profil.Role.OP_DISTRICT, Profil.Role.OP_PROVINCE)
@require_http_methods(["POST"])
def fiche_valider(request, pk):
    """Valide une fiche au palier correspondant au rôle connecté.

    Le code officiel de la paroisse est généré uniquement au moment où
    l'OP PROVINCE fait passer la fiche au statut final VALIDEE.
    Si la génération du code échoue, la validation finale est annulée
    par transaction : la fiche reste en attente OP PROVINCE.
    """
    fiche = get_object_or_404(FicheParoisse, pk=pk)
    role = get_role(request.user)
    profil = getattr(request.user, "profil", None)

    if not peut_valider_fiche(request.user, fiche):
        messages.error(request, "Cette fiche n'est pas en attente de votre validation ou se trouve hors de votre périmètre.")
        return redirect("recensement:fiche_a_valider")

    if role == Profil.Role.OP_DISTRICT:
        fiche.statut_validation = FicheParoisse.StatutValidation.ATTENTE_MANAGER
        fiche.valide_par_superviseur = request.user
        fiche.date_validation_superviseur = timezone.now()
        fiche.save(update_fields=[
            "statut_validation",
            "valide_par_superviseur",
            "date_validation_superviseur",
        ])
        messages.success(
            request,
            f"Fiche « {fiche.nom_paroisse} » validée, transmise à l'OP PROVINCE.",
        )

    elif role == Profil.Role.OP_PROVINCE:
        try:
            with transaction.atomic():
                fiche.statut_validation = FicheParoisse.StatutValidation.VALIDEE
                fiche.valide_par_manager = request.user
                fiche.date_validation_manager = timezone.now()
                fiche.save(update_fields=[
                    "statut_validation",
                    "valide_par_manager",
                    "date_validation_manager",
                ])
                code = generer_code_paroisse(fiche, genere_par=request.user)

            messages.success(
                request,
                f"Fiche « {fiche.nom_paroisse} » validée définitivement. "
                f"Code officiel généré : {code}.",
            )
        except ValueError as exc:
            messages.error(
                request,
                "La validation finale n'a pas été enregistrée, car le code officiel "
                f"n'a pas pu être généré : {exc}",
            )

    return redirect("recensement:fiche_a_valider")


# ---------------------------------------------------------------------------
# Tableau de bord
# ---------------------------------------------------------------------------

@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def dashboard(request):
    total_general = FicheParoisse.objects.count()
    total_valide = FicheParoisse.objects.filter(
        statut_validation=FicheParoisse.StatutValidation.VALIDEE
    ).count()
    total_attente_superviseur = FicheParoisse.objects.filter(
        statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR
    ).count()
    total_attente_manager = FicheParoisse.objects.filter(
        statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER
    ).count()

    par_district = list(
        FicheParoisse.objects
        .filter(statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR)
        .values("district_id", "district__nom")
        .annotate(nb=Count("id"))
        .order_by("-nb")
    )
    for ligne in par_district:
        responsables = Profil.objects.filter(
            role=Profil.Role.OP_DISTRICT, district_id=ligne["district_id"]
        ).select_related("user")
        ligne["responsables"] = [
            (p.user.get_full_name() or p.user.get_username()) for p in responsables
        ] or ["Aucun OP DISTRICT assigné"]

    par_province = list(
        FicheParoisse.objects
        .filter(statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER)
        .values("province_id", "province__nom")
        .annotate(nb=Count("id"))
        .order_by("-nb")
    )
    for ligne in par_province:
        responsables = Profil.objects.filter(
            role=Profil.Role.OP_PROVINCE, province_id=ligne["province_id"]
        ).select_related("user")
        ligne["responsables"] = [
            (p.user.get_full_name() or p.user.get_username()) for p in responsables
        ] or ["Aucun OP PROVINCE assigné"]

    context = {
        "total_general": total_general,
        "total_valide": total_valide,
        "total_attente_superviseur": total_attente_superviseur,
        "total_attente_manager": total_attente_manager,
        "par_district": par_district,
        "par_province": par_province,
    }
    return render(request, "recensement/dashboard.html", context)


@login_required
@role_required(Profil.Role.SUPER_ADMIN)
@require_GET
def suivi_modifications(request):
    historique = HistoriqueModification.objects.select_related(
        "fiche", "modifie_par"
    ).order_by("-date_modification")[:500]
    return render(request, "recensement/suivi_modifications.html", {"historique": historique})


# ---------------------------------------------------------------------------
# Carte des paroisses
# ---------------------------------------------------------------------------

@login_required
@role_required(Profil.Role.SUPER_ADMIN, Profil.Role.OP_PROVINCE, Profil.Role.OP_DISTRICT)
@require_GET
def carte_paroisses(request):
    return render(request, "recensement/carte.html")


@login_required
@role_required(Profil.Role.SUPER_ADMIN, Profil.Role.OP_PROVINCE, Profil.Role.OP_DISTRICT)
@require_GET
def fiches_geojson(request):
    fiches = (
        _fiches_visibles_pour(request.user)
        .filter(latitude__isnull=False, longitude__isnull=False)
        .select_related("region", "province", "district", "zone", "cree_par")
    )
    role = get_role(request.user)

    statut_filtre = request.GET.get("statut", "")
    if role == Profil.Role.SUPER_ADMIN and statut_filtre != "tous":
        fiches = fiches.filter(statut_validation=FicheParoisse.StatutValidation.VALIDEE)

    features = []
    for f in fiches:
        agent = f.cree_par.get_full_name() or f.cree_par.get_username() if f.cree_par else "—"
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [float(f.longitude), float(f.latitude)]},
            "properties": {
                "id": f.pk,
                "nom": f.nom_paroisse,
                "localite": f.localite,
                "zone": f.zone.nom,
                "districtId": f.district_id,
                "district": f.district.nom,
                "province": f.province.nom,
                "region": f.region.nom,
                "chargeParoisse": f.parish_shepherd,
                "agent": agent,
                "statutCode": f.statut_validation,
                "statutLabel": f.get_statut_validation_display(),
                "precisionGps": f.precision_gps,
                "url": reverse("recensement:fiche_detail", args=[f.pk]),
            },
        })

    return JsonResponse({"type": "FeatureCollection", "features": features})


# ---------------------------------------------------------------------------
# Endpoints AJAX pour les listes déroulantes en cascade
# ---------------------------------------------------------------------------

@login_required
@require_GET
def ajax_provinces(request, region_id):
    qs = Province.objects.filter(region_id=region_id)
    role = get_role(request.user)
    profil = getattr(request.user, "profil", None)
    if role != Profil.Role.SUPER_ADMIN:
        if role == Profil.Role.OP_PROVINCE and profil and profil.province_id:
            qs = qs.filter(pk=profil.province_id)
        else:
            zone_ids = zones_autorisees(request.user) or set()
            qs = qs.filter(districts__zones__id__in=zone_ids).distinct()
    return JsonResponse({"results": list(qs.order_by("nom").values("id", "nom"))})


@login_required
@require_GET
def ajax_districts(request, province_id):
    qs = District.objects.filter(province_id=province_id)
    district_ids = districts_autorises(request.user)
    if district_ids is not None:
        qs = qs.filter(pk__in=district_ids)
    return JsonResponse({"results": list(qs.order_by("nom").values("id", "nom"))})


@login_required
@require_GET
def ajax_zones(request, district_id):
    qs = Zone.objects.filter(district_id=district_id)
    zone_ids = zones_autorisees(request.user)
    if zone_ids is not None:
        qs = qs.filter(pk__in=zone_ids)
    return JsonResponse({"results": list(qs.order_by("nom").values("id", "nom"))})


@login_required
@require_GET
def ajax_villages(request, zone_id):
    zone_ids = zones_autorisees(request.user)
    if zone_ids is not None and zone_id not in zone_ids:
        return JsonResponse({"results": []}, status=403)
    villages = Village.objects.filter(zone_id=zone_id).order_by("nom").values("id", "nom")
    return JsonResponse({"results": list(villages)})


# ---------------------------------------------------------------------------
# Gestion des comptes utilisateurs
# ---------------------------------------------------------------------------

def _utilisateurs_visibles_pour(user):
    """Retourne le queryset des utilisateurs que le créateur connecté peut voir.

    - super_admin  : tous les utilisateurs.
    - op_province  : utilisateurs de sa province.
    - op_district  : utilisateurs de son district.
    - op_zone      : utilisateurs de sa zone.
    - agent        : aucun (redirection 403).
    """
    role = get_role(user)
    qs = User.objects.select_related(
        "profil", "profil__region", "profil__province",
        "profil__district", "profil__zone", "profil__cree_par",
    ).order_by("username")

    if role == Profil.Role.SUPER_ADMIN:
        return qs

    profil = getattr(user, "profil", None)
    if not profil:
        return User.objects.none()

    if role == Profil.Role.OP_PROVINCE and profil.province_id:
        return qs.filter(profil__province_id=profil.province_id)

    if role == Profil.Role.OP_DISTRICT and profil.district_id:
        return qs.filter(profil__district_id=profil.district_id)

    if role == Profil.Role.OP_ZONE and profil.zone_id:
        return qs.filter(profil__zone_id=profil.zone_id)

    return User.objects.none()


@login_required
@require_GET
def utilisateur_list(request):
    """Liste des utilisateurs — accessible aux opérateurs habilités à créer."""
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Vous n'avez pas les droits nécessaires pour accéder à cette page.")

    utilisateurs = _utilisateurs_visibles_pour(request.user)
    role = get_role(request.user)

    # Filtres disponibles pour le super admin
    filtre_role = (request.GET.get("role") or "").strip()
    filtre_region = (request.GET.get("region") or "").strip()
    filtre_province = (request.GET.get("province") or "").strip()
    filtre_district = (request.GET.get("district") or "").strip()
    filtre_zone = (request.GET.get("zone") or "").strip()

    if role == Profil.Role.SUPER_ADMIN:
        if filtre_role and filtre_role in [r.value for r in Profil.Role]:
            utilisateurs = utilisateurs.filter(profil__role=filtre_role)
        if filtre_region.isdigit():
            utilisateurs = utilisateurs.filter(profil__region_id=int(filtre_region))
        if filtre_province.isdigit():
            utilisateurs = utilisateurs.filter(profil__province_id=int(filtre_province))
        if filtre_district.isdigit():
            utilisateurs = utilisateurs.filter(profil__district_id=int(filtre_district))
        if filtre_zone.isdigit():
            utilisateurs = utilisateurs.filter(profil__zone_id=int(filtre_zone))

    return render(request, "recensement/utilisateur_list.html", {
        "utilisateurs": utilisateurs,
        "roles": Profil.Role.choices,
        "regions": Region.objects.all() if role == Profil.Role.SUPER_ADMIN else [],
        "provinces": Province.objects.all() if role == Profil.Role.SUPER_ADMIN else [],
        "filtre_role": filtre_role,
        "filtre_region": filtre_region,
        "filtre_province": filtre_province,
        "filtre_district": filtre_district,
        "filtre_zone": filtre_zone,
    })


@login_required
@require_http_methods(["GET", "POST"])
def utilisateur_create(request):
    """Création d'un utilisateur.

    Accès : tout utilisateur autorisé à créer (super_admin, op_province,
    op_district, op_zone). L'identifiant est généré automatiquement.
    Le mot de passe provisoire est affiché UNE SEULE FOIS après la création.
    """
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Vous n'avez pas les droits nécessaires.")

    if request.method == "POST":
        profil_form = ProfilForm(request.POST, createur=request.user)
        if profil_form.is_valid():
            role_cible = profil_form.cleaned_data["role"]
            region = profil_form.cleaned_data.get("region")
            province = profil_form.cleaned_data.get("province")
            district = profil_form.cleaned_data.get("district")
            zone = profil_form.cleaned_data.get("zone")

            # Vérification du périmètre (sécurité serveur, ne se base pas
            # uniquement sur ce que le formulaire propose).
            ok, msg = perimetre_creation_autorise(request.user, {
                "region_id":   region.pk if region else None,
                "province_id": province.pk if province else None,
                "district_id": district.pk if district else None,
                "zone_id":     zone.pk if zone else None,
            })
            if not ok:
                messages.error(request, msg)
                return render(request, "recensement/utilisateur_form.html", {
                    "profil_form": profil_form, "is_edit": False,
                    "regions": _regions_disponibles(request.user),
                    "provinces": _provinces_disponibles(request.user),
                })

            try:
                with transaction.atomic():
                    username = generer_identifiant(
                        role=role_cible,
                        region=region,
                        province=province,
                        district=district,
                        zone=zone,
                    )
                    mdp = generer_mot_de_passe_provisoire()
                    nouvel_utilisateur = User.objects.create_user(
                        username=username,
                        password=mdp,
                        first_name=request.POST.get("first_name", "").strip(),
                        last_name=request.POST.get("last_name", "").strip(),
                    )
                    # Le signal post_save a créé un Profil par défaut ; on le met à jour.
                    profil = nouvel_utilisateur.profil
                    profil.role = role_cible
                    profil.region = region
                    profil.province = province
                    profil.district = district
                    profil.zone = zone
                    profil.cree_par = request.user
                    profil.save()

                    # Le mot de passe provisoire est stocké dans la session
                    # pour être affiché UNE SEULE FOIS sur la page de confirmation.
                    request.session["mdp_provisoire_username"] = username
                    request.session["mdp_provisoire_valeur"] = mdp

            except ValueError as e:
                messages.error(request, f"Erreur de génération de l'identifiant : {e}")
                return render(request, "recensement/utilisateur_form.html", {
                    "profil_form": profil_form, "is_edit": False,
                    "regions": _regions_disponibles(request.user),
                    "provinces": _provinces_disponibles(request.user),
                })

            return redirect("recensement:utilisateur_created", pk=nouvel_utilisateur.pk)

        messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        profil_form = ProfilForm(createur=request.user)

    return render(request, "recensement/utilisateur_form.html", {
        "profil_form": profil_form,
        "is_edit": False,
        "regions": _regions_disponibles(request.user),
        "provinces": _provinces_disponibles(request.user),
    })


@login_required
@require_GET
def utilisateur_created(request, pk):
    """Page de confirmation après création d'un utilisateur.

    Affiche le mot de passe provisoire UNE SEULE FOIS, puis le supprime
    de la session. L'administrateur doit copier et transmettre ce mot de
    passe à l'utilisateur par un canal sécurisé.
    """
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied()

    utilisateur = get_object_or_404(User, pk=pk)

    # Récupération et suppression immédiate du mot de passe provisoire.
    mdp = request.session.pop("mdp_provisoire_valeur", None)
    mdp_username = request.session.pop("mdp_provisoire_username", None)

    # Sécurité : on ne réaffiche le mot de passe que si la session correspond
    # bien à cet utilisateur (évite qu'un autre admin accède à l'URL directement).
    if mdp_username != utilisateur.username:
        mdp = None

    return render(request, "recensement/utilisateur_created.html", {
        "utilisateur": utilisateur,
        "mdp_provisoire": mdp,
    })


@login_required
@require_http_methods(["GET", "POST"])
def utilisateur_update(request, pk):
    """Modification d'un utilisateur — accessible aux opérateurs habilités."""
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied()

    utilisateur = get_object_or_404(_utilisateurs_visibles_pour(request.user), pk=pk)
    profil, _ = Profil.objects.get_or_create(user=utilisateur)

    if request.method == "POST":
        profil_form = ProfilForm(request.POST, instance=profil, createur=request.user)
        if profil_form.is_valid():
            role_cible = profil_form.cleaned_data["role"]
            province = profil_form.cleaned_data.get("province")
            district = profil_form.cleaned_data.get("district")
            zone = profil_form.cleaned_data.get("zone")
            region = profil_form.cleaned_data.get("region")

            ok, msg = perimetre_creation_autorise(request.user, {
                "region_id":   region.pk if region else None,
                "province_id": province.pk if province else None,
                "district_id": district.pk if district else None,
                "zone_id":     zone.pk if zone else None,
            })
            if not ok:
                messages.error(request, msg)
            else:
                profil_form.save()
                utilisateur.first_name = request.POST.get("first_name", "").strip()
                utilisateur.last_name = request.POST.get("last_name", "").strip()
                utilisateur.is_active = request.POST.get("is_active") == "on"
                utilisateur.save()
                messages.success(request, "Compte mis à jour avec succès.")
                return redirect("recensement:utilisateur_list")
        messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        profil_form = ProfilForm(instance=profil, createur=request.user)

    return render(request, "recensement/utilisateur_form.html", {
        "profil_form": profil_form,
        "is_edit": True,
        "utilisateur": utilisateur,
        "regions": _regions_disponibles(request.user),
        "provinces": _provinces_disponibles(request.user),
        "province_du_district_id": profil.district.province_id if profil.district_id else None,
        "zone_du_district_id": profil.zone.district_id if profil.zone_id else None,
    })


@login_required
@require_http_methods(["GET", "POST"])
def utilisateur_reset_password(request, pk):
    """Réinitialisation du mot de passe — accessible aux opérateurs habilités."""
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied()

    utilisateur = get_object_or_404(_utilisateurs_visibles_pour(request.user), pk=pk)

    if request.method == "POST":
        form = TailwindSetPasswordForm(utilisateur, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"Mot de passe réinitialisé pour « {utilisateur.get_username()} ».")
            return redirect("recensement:utilisateur_list")
        messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = TailwindSetPasswordForm(utilisateur)

    return render(request, "recensement/utilisateur_reset_password.html", {
        "form": form, "utilisateur": utilisateur,
    })


@login_required
@require_http_methods(["POST"])
def utilisateur_toggle_actif(request, pk):
    """Activation/désactivation d'un compte — accessible aux opérateurs habilités."""
    if not peut_creer_utilisateur(request.user):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied()

    utilisateur = get_object_or_404(_utilisateurs_visibles_pour(request.user), pk=pk)
    if utilisateur == request.user:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
    else:
        utilisateur.is_active = not utilisateur.is_active
        utilisateur.save()
        etat = "réactivé" if utilisateur.is_active else "désactivé"
        messages.success(request, f"Compte « {utilisateur.get_username()} » {etat}.")
    return redirect("recensement:utilisateur_list")


@login_required
@require_http_methods(["GET", "POST"])
def utilisateur_delete(request, pk):
    """Suppression d'un compte — réservée au super admin."""
    role = get_role(request.user)
    if role != Profil.Role.SUPER_ADMIN:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied()

    utilisateur = get_object_or_404(User, pk=pk)
    if utilisateur == request.user:
        messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
        return redirect("recensement:utilisateur_list")

    if request.method == "POST":
        nom = utilisateur.get_username()
        utilisateur.delete()
        messages.success(request, f"Le compte « {nom} » a été supprimé définitivement.")
        return redirect("recensement:utilisateur_list")

    return render(request, "recensement/utilisateur_confirm_delete.html", {"utilisateur": utilisateur})


# ---------------------------------------------------------------------------
# Helpers internes : périmètres disponibles selon le créateur
# ---------------------------------------------------------------------------

def _regions_disponibles(user):
    """Régions que le créateur connecté peut sélectionner pour un nouveau compte."""
    role = get_role(user)
    if role == Profil.Role.SUPER_ADMIN:
        return Region.objects.all()
    profil = getattr(user, "profil", None)
    if profil and profil.region_id:
        return Region.objects.filter(pk=profil.region_id)
    return Region.objects.none()


def _provinces_disponibles(user):
    """Provinces que le créateur connecté peut sélectionner."""
    role = get_role(user)
    if role == Profil.Role.SUPER_ADMIN:
        return Province.objects.select_related("region").all()
    profil = getattr(user, "profil", None)
    if profil and profil.province_id:
        return Province.objects.filter(pk=profil.province_id)
    return Province.objects.none()


# ---------------------------------------------------------------------------
# Helpers export (utilisé par fiche_export_preview et fiche_export_excel)
# ---------------------------------------------------------------------------

def _fiches_export_filtrees(request):
    fiches = _fiches_visibles_pour(request.user)
    role = get_role(request.user)

    statut_filtre = request.GET.get("statut", "")
    if role == Profil.Role.SUPER_ADMIN:
        if statut_filtre == "attente_superviseur":
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_SUPERVISEUR
            )
        elif statut_filtre == "attente_manager":
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.ATTENTE_MANAGER
            )
        elif statut_filtre == "tous":
            pass
        else:
            fiches = fiches.filter(
                statut_validation=FicheParoisse.StatutValidation.VALIDEE
            )

    def valid_id(param_name):
        value = (request.GET.get(param_name) or "").strip()
        return int(value) if value.isdigit() else None

    region_id = valid_id("region")
    province_id = valid_id("province")
    district_id = valid_id("district")
    zone_id = valid_id("zone")
    paroisse = (request.GET.get("paroisse") or "").strip()[:100]

    if region_id:
        fiches = fiches.filter(region_id=region_id)
    if province_id:
        fiches = fiches.filter(province_id=province_id)
    if district_id:
        fiches = fiches.filter(district_id=district_id)
    if zone_id:
        fiches = fiches.filter(zone_id=zone_id)
    if paroisse:
        fiches = fiches.filter(nom_paroisse__icontains=paroisse)

    return fiches.select_related(
        "region", "province", "district", "zone", "village"
    ).order_by(
        "region__nom", "province__nom", "district__nom", "zone__nom", "nom_paroisse",
    )
