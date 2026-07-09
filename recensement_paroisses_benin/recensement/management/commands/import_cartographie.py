"""
Commande de gestion Django : import du référentiel géo-ecclésial du Bénin
(Région > Province > District > Zone > Village) à partir du fichier Excel
de cartographie fourni par le diocèse.

Usage :
    python manage.py import_cartographie
    python manage.py import_cartographie --file /chemin/vers/fichier.xlsx
    python manage.py import_cartographie --sheet "Cartographie avec villes " --dry-run

Le fichier Excel attendu a la structure suivante (une colonne par niveau,
lecture "en cascade" de haut en bas — chaque ligne ne renseigne qu'UN seul
niveau, le reste étant hérité du dernier niveau parent rencontré) :

    Col A : Nom de la région        ex. "BORGOU-ALIBORI (2ème Région ecclésiale)"
    Col B : "↳ Province ecclésiale de X [Y]"
    Col C : "↳ District ecclésial de X (Y)"
    Col D : "• Zone ecclésiale de X"  ou  "• Site de ..."
    Col E : villages/quartiers séparés par ";" ou "," (et éventuellement " et "
            avant le dernier élément) — uniquement présent dans la feuille
            "Cartographie avec villes ".
    Col F : nombre de villages (contrôle de cohérence, informatif)
"""

import re

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from recensement.models import District, Province, Region, Village, Zone

DEFAULT_SHEET = "Cartographie avec villes "
DEFAULT_FILE = settings.BASE_DIR / "recensement" / "data" / "cartographie_benin.xlsx"


def clean_region(text):
    """'BORGOU-ALIBORI (2ème Région ecclésiale)' -> ('BORGOU-ALIBORI', 2)"""
    text = text.strip()
    m = re.search(r"\((\d+)", text)
    ordre = int(m.group(1)) if m else 0
    nom = re.sub(r"\(.*?\)", "", text).strip()
    nom = re.sub(r"\s{2,}", " ", nom)
    return nom, ordre


def _strip_prefix(text, prefixes):
    text = text.strip().lstrip("↳").strip()
    low = text.lower()
    for p in prefixes:
        if low.startswith(p.lower()):
            text = text[len(p):].strip()
            break
    return text


def clean_province(text):
    """'↳ Province ecclésiale de Ouémé [Ouémé]' -> 'Ouémé'"""
    text = _strip_prefix(
        text, ["Province ecclésiale de ", "Province ecclésiale d'", "Province ecclésiale "]
    )
    text = re.sub(r"\s*\[.*?\]\s*$", "", text).strip()
    return text


def clean_district(text):
    """'↳ District ecclésial de Porto-Novo (Porto-Novo)' -> 'Porto-Novo'"""
    text = _strip_prefix(
        text, ["District ecclésial de ", "District ecclésial d'", "District ecclésial "]
    )
    text = re.sub(r"\s*\(.*?\)\s*$", "", text).strip()
    return text


def clean_zone(text):
    """'• Zone ecclésiale de Banikoara' -> 'Banikoara'
    '• Site de la nativité de SÈMÈ PLAGE' -> 'Site de la nativité de SÈMÈ PLAGE'
    (les 'Sites particuliers' gardent leur nom complet, seul le puce est retiré)."""
    text = text.strip().lstrip("•").strip()
    text = _strip_prefix(
        text, ["Zone ecclésiale de ", "Zone ecclésiale d'", "Zone ecclésiale "]
    )
    return text


def split_villages(raw):
    """Découpe la cellule 'Villages / quartiers' en une liste de noms propres.
    Gère les séparateurs ';' et ',' ainsi que le ' et ' final français."""
    if not raw:
        return []
    text = str(raw).strip()
    text = re.sub(r"\s+et\s+", ", ", text, flags=re.IGNORECASE)
    parts = re.split(r"[;,]", text)
    out = []
    for p in parts:
        p = p.strip().strip(".").strip().strip('"').strip("'").strip()
        if p:
            out.append(p)
    return out


class Command(BaseCommand):
    help = (
        "Importe le référentiel géo-ecclésial du Bénin (région > province > "
        "district > zone > village) depuis le fichier Excel de cartographie."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", type=str, default=str(DEFAULT_FILE),
            help="Chemin vers le fichier Excel (par défaut : recensement/data/cartographie_benin.xlsx)",
        )
        parser.add_argument(
            "--sheet", type=str, default=DEFAULT_SHEET,
            help="Nom de la feuille à lire (par défaut : la feuille avec les villages).",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Analyse le fichier et affiche le résumé sans rien écrire en base.",
        )

    def handle(self, *args, **options):
        filepath = options["file"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fichier introuvable : {filepath}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(
                f"Feuille '{sheet_name}' introuvable. Feuilles disponibles : {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        stats = {
            "regions": 0, "provinces": 0, "districts": 0, "zones": 0, "villages": 0,
            "lignes_ignorees": 0,
        }

        current_region = None
        current_province = None
        current_district = None

        with transaction.atomic():
            sid = transaction.savepoint()

            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
                # On ne lit que les 6 premières colonnes utiles (A-F)
                a, b, c, d, e, f = (list(row) + [None] * 6)[:6]

                try:
                    # --- Niveau Région ---
                    if a and "CARTOGRAPHIE" not in str(a).upper():
                        nom, ordre = clean_region(str(a))
                        if not nom:
                            stats["lignes_ignorees"] += 1
                            continue
                        region_obj, created = Region.objects.get_or_create(
                            nom=nom, defaults={"ordre": ordre}
                        )
                        if created:
                            stats["regions"] += 1
                        current_region = region_obj
                        current_province = None
                        current_district = None
                        continue

                    # --- Niveau Province ---
                    if b:
                        if current_region is None:
                            stats["lignes_ignorees"] += 1
                            continue
                        nom = clean_province(str(b))
                        if not nom:
                            stats["lignes_ignorees"] += 1
                            continue
                        province_obj, created = Province.objects.get_or_create(
                            region=current_region, nom=nom
                        )
                        if created:
                            stats["provinces"] += 1
                        current_province = province_obj
                        current_district = None
                        continue

                    # --- Niveau District ---
                    if c:
                        if current_province is None:
                            stats["lignes_ignorees"] += 1
                            continue
                        nom = clean_district(str(c))
                        if not nom:
                            stats["lignes_ignorees"] += 1
                            continue
                        district_obj, created = District.objects.get_or_create(
                            province=current_province, nom=nom
                        )
                        if created:
                            stats["districts"] += 1
                        current_district = district_obj
                        continue

                    # --- Niveau Zone (+ villages associés) ---
                    if d:
                        if current_district is None:
                            stats["lignes_ignorees"] += 1
                            continue
                        nom = clean_zone(str(d))
                        if not nom:
                            stats["lignes_ignorees"] += 1
                            continue
                        zone_obj, created = Zone.objects.get_or_create(
                            district=current_district, nom=nom
                        )
                        if created:
                            stats["zones"] += 1

                        for village_nom in split_villages(e):
                            village_nom = village_nom[:200]
                            _, v_created = Village.objects.get_or_create(
                                zone=zone_obj, nom=village_nom
                            )
                            if v_created:
                                stats["villages"] += 1
                        continue

                except Exception as exc:  # pragma: no cover - garde-fou d'import
                    self.stderr.write(
                        self.style.WARNING(f"Ligne {i} ignorée ({exc}) : {row}")
                    )
                    stats["lignes_ignorees"] += 1

            if dry_run:
                transaction.savepoint_rollback(sid)
                self.stdout.write(self.style.WARNING(
                    "--dry-run : aucune donnée n'a été écrite en base."
                ))
            else:
                transaction.savepoint_commit(sid)

        self.stdout.write(self.style.SUCCESS(
            "Import terminé :\n"
            f"  Régions ajoutées   : {stats['regions']}\n"
            f"  Provinces ajoutées : {stats['provinces']}\n"
            f"  Districts ajoutés  : {stats['districts']}\n"
            f"  Zones ajoutées     : {stats['zones']}\n"
            f"  Villages ajoutés   : {stats['villages']}\n"
            f"  Lignes ignorées    : {stats['lignes_ignorees']}"
        ))
